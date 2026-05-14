# -*- coding: utf-8 -*-
import base64
import io
import json
import logging
import sys

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

COLUMNS = [
    'code', 'name', 'category', 'trade', 'uom', 'unit_rate',
    'currency', 'valid_from', 'valid_to', 'notes', 'active', 'company',
]
SCOPES = ['https://www.googleapis.com/auth/drive.file']


def _to_bool(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ('1', 'true', 'yes', 'y', 't')


class BuruujRate(models.Model):
    _inherit = 'buruuj.rate'

    def action_bulk_activate(self):
        self.write({'active': True})
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def action_bulk_deactivate(self):
        self.write({'active': False})
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def action_gdrive_push(self):
        self.env['buruuj.rate']._gdrive_push()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Pushed to Google Drive'),
                'message': _('Master rates exported successfully.'),
                'type': 'success',
            },
        }

    def action_gdrive_pull(self):
        created, updated = self.env['buruuj.rate']._gdrive_pull()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Pulled from Google Drive'),
                'message': _('%(c)d created, %(u)d updated.', c=created, u=updated),
                'type': 'success',
                'next': {'type': 'ir.actions.client', 'tag': 'reload'},
            },
        }

    @api.model
    def _gdrive_get_service(self):
        try:
            from googleapiclient.discovery import build
            from google.oauth2 import service_account
        except ImportError as e:
            raise UserError(_(
                'Google API libraries not installed.\n'
                'Run: "%s" -m pip install google-api-python-client google-auth\n\n%s',
                sys.executable, e))

        cp = self.env['ir.config_parameter'].sudo()
        creds_b64 = cp.get_param('buruuj_tendering_gdrive.service_account_json')
        if not creds_b64:
            raise UserError(_('No service account credentials configured. '
                              'Use "Drive Setup" on the Master Rates list.'))
        try:
            creds_json = json.loads(base64.b64decode(creds_b64).decode('utf-8'))
        except Exception as e:
            raise UserError(_('Could not parse service account JSON: %s', e))

        creds = service_account.Credentials.from_service_account_info(
            creds_json, scopes=SCOPES)
        return build('drive', 'v3', credentials=creds, cache_discovery=False)

    @api.model
    def _gdrive_get_config(self):
        cp = self.env['ir.config_parameter'].sudo()
        folder_id = cp.get_param('buruuj_tendering_gdrive.folder_id')
        filename = cp.get_param(
            'buruuj_tendering_gdrive.filename') or 'buruuj_master_rates.xlsx'
        if not folder_id:
            raise UserError(_('No Drive folder configured. '
                              'Use "Drive Setup" on the Master Rates list.'))
        return folder_id, filename

    @api.model
    def _gdrive_find_file(self, service, folder_id, filename):
        safe_name = filename.replace("'", "\\'")
        q = "name = '%s' and '%s' in parents and trashed = false" % (safe_name, folder_id)
        result = service.files().list(
            q=q, fields='files(id, name, modifiedTime)', pageSize=1).execute()
        files = result.get('files', [])
        return files[0] if files else None

    @api.model
    def _gdrive_push(self):
        from googleapiclient.http import MediaIoBaseUpload
        import openpyxl

        folder_id, filename = self._gdrive_get_config()
        service = self._gdrive_get_service()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'master_rates'
        ws.append(COLUMNS)
        for r in self.search([]):
            ws.append([
                r.code or '',
                r.name or '',
                r.category or '',
                r.trade_id.name or '',
                r.uom_id.name or '',
                float(r.unit_rate or 0.0),
                r.currency_id.name or '',
                r.valid_from.isoformat() if r.valid_from else '',
                r.valid_to.isoformat() if r.valid_to else '',
                r.notes or '',
                bool(r.active),
                r.company_id.name or '',
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        media = MediaIoBaseUpload(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=False)

        existing = self._gdrive_find_file(service, folder_id, filename)
        if existing:
            service.files().update(
                fileId=existing['id'], media_body=media).execute()
            _logger.info('Buruuj rate push: updated %s (id=%s)',
                         filename, existing['id'])
        else:
            metadata = {'name': filename, 'parents': [folder_id]}
            f = service.files().create(
                body=metadata, media_body=media, fields='id').execute()
            _logger.info('Buruuj rate push: created %s (id=%s)',
                         filename, f.get('id'))

    @api.model
    def _gdrive_pull(self):
        from googleapiclient.http import MediaIoBaseDownload
        import openpyxl

        folder_id, filename = self._gdrive_get_config()
        service = self._gdrive_get_service()

        existing = self._gdrive_find_file(service, folder_id, filename)
        if not existing:
            raise UserError(_(
                'No %(name)s found in the configured Drive folder. '
                'Push first, or upload the file manually.', name=filename))

        request = service.files().get_media(fileId=existing['id'])
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)

        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            raise UserError(_('Drive file is empty.'))

        idx = {col: header.index(col) for col in COLUMNS if col in header}
        if 'code' not in idx:
            raise UserError(_('Drive file is missing required column "code". '
                              'Header found: %s', header))

        Trade = self.env['buruuj.trade']
        Uom = self.env['uom.uom']
        Currency = self.env['res.currency']
        Company = self.env['res.company']

        created = updated = 0
        for row in rows_iter:
            code = row[idx['code']]
            if not code:
                continue
            code = str(code).strip()

            company = self.env.company
            if 'company' in idx and row[idx['company']]:
                hit = Company.search(
                    [('name', '=', str(row[idx['company']]).strip())], limit=1)
                if hit:
                    company = hit

            vals = {
                'code': code,
                'name': str(row[idx['name']] or '').strip() if 'name' in idx else '',
                'category': (str(row[idx['category']]).strip()
                             if 'category' in idx and row[idx['category']]
                             else 'material'),
                'unit_rate': float(row[idx['unit_rate']] or 0.0)
                             if 'unit_rate' in idx else 0.0,
                'active': _to_bool(row[idx['active']]) if 'active' in idx else True,
                'company_id': company.id,
            }
            if 'notes' in idx and row[idx['notes']]:
                vals['notes'] = str(row[idx['notes']])
            if 'valid_from' in idx and row[idx['valid_from']]:
                vals['valid_from'] = row[idx['valid_from']]
            if 'valid_to' in idx and row[idx['valid_to']]:
                vals['valid_to'] = row[idx['valid_to']]
            if 'trade' in idx and row[idx['trade']]:
                trade = Trade.search(
                    [('name', '=', str(row[idx['trade']]).strip())], limit=1)
                if trade:
                    vals['trade_id'] = trade.id
            if 'uom' in idx and row[idx['uom']]:
                uom = Uom.search(
                    [('name', '=', str(row[idx['uom']]).strip())], limit=1)
                if uom:
                    vals['uom_id'] = uom.id
            if 'currency' in idx and row[idx['currency']]:
                cur = Currency.search(
                    [('name', '=', str(row[idx['currency']]).strip())], limit=1)
                if cur:
                    vals['currency_id'] = cur.id

            existing_rate = self.search(
                [('code', '=', code), ('company_id', '=', company.id)], limit=1)
            if existing_rate:
                existing_rate.write(vals)
                updated += 1
            else:
                if 'uom_id' not in vals:
                    raise UserError(_(
                        'Cannot create rate %(code)s — no matching unit of measure '
                        'for "%(uom)s" found in Odoo.',
                        code=code,
                        uom=(row[idx['uom']] if 'uom' in idx else '')))
                self.create(vals)
                created += 1

        _logger.info('Buruuj rate pull: %d created, %d updated', created, updated)
        return created, updated

    @api.model
    def _cron_gdrive_pull(self):
        try:
            self._gdrive_pull()
        except Exception:
            _logger.exception('Buruuj rate cron pull failed')
