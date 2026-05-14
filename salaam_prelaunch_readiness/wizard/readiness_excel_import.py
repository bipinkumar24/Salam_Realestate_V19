# -*- coding: utf-8 -*-
"""
One-shot Excel import wizard.

For tenants migrating from the source spreadsheet with partial progress
already entered (Status, % Complete, Owner, Due Date, Notes). Maps
spreadsheet rows to seeded items by category code + sequence number.
Idempotent: running it twice with the same file produces the same end
state.
"""
import base64
import io
import logging

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# Map spreadsheet sheet name prefix -> category code
SHEET_PREFIX_TO_CODE = {
    "A.": "A", "B.": "B", "C.": "C", "D.": "D", "E.": "E",
    "F.": "F", "G.": "G", "H.": "H", "I.": "I", "J.": "J",
    "K.": "K", "L.": "L", "M.": "M", "N.": "N",
}

# Map spreadsheet status text -> internal selection key
STATUS_MAP = {
    "not started": "not_started",
    "in progress": "in_progress",
    "done": "done",
    "blocked": "blocked",
    "n/a": "na",
    "na": "na",
}


class ReadinessExcelImport(models.TransientModel):
    _name = "salaam.readiness.excel.import"
    _description = "Salaam Readiness - Excel Import Wizard"

    tracker_id = fields.Many2one(
        "salaam.readiness.tracker",
        string="Target Tracker",
        required=True,
    )
    excel_file = fields.Binary(string="Excel File (.xlsx)", required=True)
    excel_filename = fields.Char(string="Filename")
    update_owner = fields.Boolean(
        string="Match Owners by Email",
        default=True,
        help="If unticked, the Owner column in the spreadsheet is ignored. "
             "If ticked, the wizard will look up Odoo users whose email "
             "matches the Owner cell text (case-insensitive).",
    )
    update_evidence_link = fields.Boolean(
        string="Import Evidence Link Column",
        default=True,
    )

    result_log = fields.Text(string="Import Log", readonly=True)
    matched_count = fields.Integer(string="Items Updated", readonly=True)
    unmatched_count = fields.Integer(string="Rows Unmatched", readonly=True)

    # ------------------------------------------------------------------
    def _parse_status(self, raw):
        if not raw:
            return None
        return STATUS_MAP.get(str(raw).strip().lower())

    def _parse_pct(self, raw):
        if raw is None or raw == "":
            return None
        try:
            v = float(raw)
        except (TypeError, ValueError):
            return None
        # Accept either 0-1 (Excel decimal %) or 0-100
        if 0.0 <= v <= 1.0:
            v = v * 100.0
        return max(0.0, min(100.0, v))

    def _resolve_owner(self, raw):
        if not raw or not self.update_owner:
            return False
        text = str(raw).strip()
        if not text:
            return False
        # Try email match first, then name
        user = self.env["res.users"].search(
            [("login", "=ilike", text)], limit=1
        )
        if not user:
            user = self.env["res.users"].search(
                [("email", "=ilike", text)], limit=1
            )
        if not user:
            user = self.env["res.users"].search(
                [("name", "=ilike", text)], limit=1
            )
        return user.id if user else False

    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Please upload an .xlsx file."))

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_(
                "The Python package 'openpyxl' is required for this "
                "import. Install with: pip install openpyxl"
            ))

        try:
            data = base64.b64decode(self.excel_file)
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(_("Could not read the Excel file: %s") % e)

        log_lines = []
        matched = 0
        unmatched = 0

        Item = self.env["salaam.readiness.item"]

        for sheet_name in wb.sheetnames:
            # Find which category this sheet belongs to
            cat_code = None
            for prefix, code in SHEET_PREFIX_TO_CODE.items():
                if sheet_name.startswith(prefix):
                    cat_code = code
                    break
            if not cat_code:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Locate the data header row (first cell == "#", second has "Item")
            header_idx = None
            for idx, row in enumerate(rows):
                if row and row[0] == "#" and row[1] and "Item" in str(row[1]):
                    header_idx = idx
                    break
            if header_idx is None:
                log_lines.append(f"[{sheet_name}] header row not found - skipped")
                continue

            for row in rows[header_idx + 1:]:
                if not row or row[0] is None:
                    continue
                try:
                    seq = int(row[0])
                except (TypeError, ValueError):
                    continue

                # Columns: # | Item | Gate Type | Owner | Due Date | Status | % Complete | Evidence Link | Notes
                _gate = row[2] if len(row) > 2 else None
                owner_raw = row[3] if len(row) > 3 else None
                due_raw = row[4] if len(row) > 4 else None
                status_raw = row[5] if len(row) > 5 else None
                pct_raw = row[6] if len(row) > 6 else None
                ev_raw = row[7] if len(row) > 7 else None
                notes_raw = row[8] if len(row) > 8 else None

                item = Item.search([
                    ("tracker_id", "=", self.tracker_id.id),
                    ("category_code", "=", cat_code),
                    ("sequence", "=", seq),
                ], limit=1)
                if not item:
                    unmatched += 1
                    log_lines.append(
                        f"[{cat_code}.{seq}] no matching seeded item - row skipped"
                    )
                    continue

                vals = {}
                status = self._parse_status(status_raw)
                if status:
                    vals["status"] = status
                pct = self._parse_pct(pct_raw)
                if pct is not None:
                    vals["pct_complete"] = pct
                owner_id = self._resolve_owner(owner_raw)
                if owner_id:
                    vals["owner_id"] = owner_id
                if due_raw:
                    try:
                        # openpyxl returns datetime if cell is date-typed
                        if hasattr(due_raw, "date"):
                            vals["due_date"] = due_raw.date()
                        else:
                            vals["due_date"] = fields.Date.to_date(str(due_raw))
                    except Exception:
                        pass
                if ev_raw and self.update_evidence_link:
                    vals["evidence_url"] = str(ev_raw).strip()
                if notes_raw:
                    vals["notes"] = str(notes_raw).strip()

                if vals:
                    item.write(vals)
                    matched += 1

        wb.close()

        self.result_log = "\n".join(log_lines) if log_lines else _("All rows matched cleanly.")
        self.matched_count = matched
        self.unmatched_count = unmatched

        # Force tracker recompute
        self.tracker_id._compute_aggregate()

        self.tracker_id.message_post(
            body=_("Excel import complete: %(m)d items updated, "
                   "%(u)d rows unmatched.") % {"m": matched, "u": unmatched},
        )

        return {
            "type": "ir.actions.act_window",
            "name": _("Import Result"),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": {"form_view_initial_mode": "readonly"},
        }
